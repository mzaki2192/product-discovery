"""
Fix col B, C and I in Epic Breakdown sheet:
- Col B header (row 3): add dark fill matching other headers
- Col B data: fix font alpha (000563C1 → FF0563C1), add hyperlinks for MFT-XXXX values
- Col C (Phase): restore phase-based fill color + white text (was overwritten with plain banding)
- Col I (Desc Quality): restore quality-based fill color + matching text
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

XLSX = "/Users/zaki/Downloads/Product discovery/Tech Debt/outputs/tech_debt_report.xlsx"
JIRA_BASE = "https://tabby.atlassian.net/browse/"

wb = openpyxl.load_workbook(XLSX)
ws = wb["Epic Breakdown"]

DATA_START = 4
HEADER_ROW = 3

# ---- Phase fill (same as col A) ----
PHASE_FILL = {
    "Phase 0": "FF2C3E50",
    "Phase 1": "FF1A5276",
    "Phase 2": "FF1E8449",
    "Phase 3": "FF6C3483",
    "Various": "FF2C3E50",
}

# ---- Desc Quality fill + text color ----
QUALITY_FILL = {
    "✓ Complete":       ("FF375623", "FFFFFFFF"),  # dark green, white text
    "OK":               ("FF375623", "FFFFFFFF"),  # dark green, white text
    "Partial (1 missing)": ("FF7F6000", "FFFFFFFF"),  # dark amber, white text
    "Partial (2 missing)": ("FF833C00", "FFFFFFFF"),  # dark orange, white text
    "✗ All missing":    ("FF9C0006", "FFFFFFFF"),  # dark red, white text
}
QUALITY_DEFAULT = ("FF7F6000", "FFFFFFFF")  # amber fallback

def make_fill(hex8):
    return PatternFill("solid", fgColor=hex8)

def thin_border():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

fixed_b_header = 0
fixed_b_data = 0
fixed_c = 0
fixed_i = 0

# ---- Fix col B header (row 3) ----
cell = ws.cell(HEADER_ROW, 2)
if cell.fill.fgColor.rgb != "FF1A1A2E":
    cell.fill = make_fill("FF1A1A2E")
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()
    fixed_b_header = 1
    print(f"  Fixed col B header (row {HEADER_ROW})")

# ---- Fix col B, C, I data rows ----
for row in range(DATA_START, ws.max_row + 1):
    phase_val = str(ws.cell(row, 3).value or "").strip()
    key_val   = str(ws.cell(row, 4).value or "").strip()
    if not key_val:
        continue

    # Determine row banding position (1-based relative to data start)
    pos = row - DATA_START + 1
    band_hex = "FFFAFAFA" if pos % 2 == 1 else "FFFFFFFF"

    # ---- Col B (Epic Link) ----
    cell_b = ws.cell(row, 2)
    b_val = str(cell_b.value or "").strip()
    if b_val and b_val != "—":
        # MFT key — hyperlink + opaque blue font
        cell_b.hyperlink = JIRA_BASE + b_val
        cell_b.font = Font(name="Calibri", size=11, color="FF0563C1", underline="single")
    else:
        # "—" or empty — plain dark text, no hyperlink
        cell_b.hyperlink = None
        cell_b.font = Font(name="Calibri", size=11, color="FF000000")
    cell_b.fill = make_fill(band_hex)
    cell_b.alignment = Alignment(horizontal="center", vertical="center")
    cell_b.border = thin_border()
    fixed_b_data += 1

    # ---- Col C (Phase) ----
    cell_c = ws.cell(row, 3)
    phase_fill_hex = PHASE_FILL.get(phase_val, "FF2C3E50")
    cell_c.fill = make_fill(phase_fill_hex)
    cell_c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    cell_c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_c.border = thin_border()
    fixed_c += 1

    # ---- Col I (Desc Quality) ----
    cell_i = ws.cell(row, 9)
    i_val = str(cell_i.value or "").strip()
    fill_hex, text_hex = QUALITY_FILL.get(i_val, QUALITY_DEFAULT)
    cell_i.fill = make_fill(fill_hex)
    cell_i.font = Font(name="Calibri", size=10, bold=True, color=text_hex)
    cell_i.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_i.border = thin_border()
    fixed_i += 1

wb.save(XLSX)
print(f"\nDone. Saved: {XLSX}")
print(f"  Col B header fixed : {fixed_b_header}")
print(f"  Col B data fixed   : {fixed_b_data} rows")
print(f"  Col C fixed        : {fixed_c} rows")
print(f"  Col I fixed        : {fixed_i} rows")
