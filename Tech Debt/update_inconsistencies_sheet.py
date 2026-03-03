"""
Rewrites the Inconsistencies sheet in tech_debt_report.xlsx using current Jira data.
- Sections 1 & 3 (actions already applied): collapsed to resolved summaries
- Section 2 (Excel not in Jira): updated with current Jira status per fetch
- Section 4 (verdict conflicts): updated with current Jira status and revised recommendations
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX = "/Users/zaki/Downloads/Product discovery/Tech Debt/outputs/tech_debt_report.xlsx"

wb = openpyxl.load_workbook(XLSX)
ws = wb["Inconsistencies"]

# Unmerge all merged cells first, then clear
for merge in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(merge))

for row in ws.iter_rows():
    for cell in row:
        cell.value = None
        cell.font = Font()
        cell.fill = PatternFill()
        cell.alignment = Alignment()
        cell.border = Border()
        cell.hyperlink = None

# ---- Style helpers ----
def title_font(size=13):
    return Font(name="Calibri", bold=True, size=size)

def header_font():
    return Font(name="Calibri", bold=True, size=10, color="FFFFFF")

def body_font(bold=False, color="000000"):
    return Font(name="Calibri", size=10, bold=bold, color=color)

def section_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def wrap_align(horizontal="left"):
    return Alignment(horizontal=horizontal, vertical="top", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def write(row, col, value, font=None, fill=None, align=None, border=None):
    cell = ws.cell(row, col)
    cell.value = value
    if font:   cell.font = font
    if fill:   cell.fill = fill
    if align:  cell.alignment = align
    if border: cell.border = border
    return cell

# ---- Column widths ----
col_widths = [14, 18, 45, 16, 24, 35, 45, 30]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ---- Row heights (set later) ----
ROW_H_TITLE = 20
ROW_H_HDR   = 15
ROW_H_DATA  = 30
ROW_H_SECTION = 18

r = 1  # current row pointer

# ==== TITLE ====
ws.merge_cells(f"A{r}:H{r}")
write(r, 1, "🔍  Inconsistencies Report — Capital Tech Debt (MFT Stability Board)",
      font=title_font(13), align=Alignment(horizontal="left", vertical="center"))
ws.row_dimensions[r].height = ROW_H_TITLE
r += 1

ws.merge_cells(f"A{r}:H{r}")
write(r, 1,
      "Updated: 2026-03-01  |  Source of Truth: Jira MFT  |  Prior sync: 2026-02-28  |  Sections 1 & 3 resolved — see archive note",
      font=body_font(color="666666"), align=Alignment(horizontal="left", vertical="center"))
ws.row_dimensions[r].height = 14
r += 2

# ==== SECTION 1 — RESOLVED ====
BLUE_DARK  = "1F4E79"
BLUE_MID   = "2E75B6"
GREEN_DARK = "1E5C2E"
GREEN_FILL = "E2EFDA"
AMBER_FILL = "FFF2CC"
RED_FILL   = "FCE4D6"
GREY_FILL  = "F2F2F2"
HEADER_FILL_BLUE  = section_fill(BLUE_MID)
HEADER_FILL_GREEN = section_fill("375623")
HEADER_FILL_AMBER = section_fill("7F6000")
HEADER_FILL_RED   = section_fill("833C00")

ws.merge_cells(f"A{r}:H{r}")
write(r, 1, "✅  SECTION 1 — In Jira but NOT in Excel  →  RESOLVED (12 tasks added to Epic Breakdown on 2026-02-28)",
      font=Font(name="Calibri", bold=True, size=11, color="FFFFFF"),
      fill=section_fill("375623"),
      align=Alignment(horizontal="left", vertical="center", wrap_text=True))
ws.row_dimensions[r].height = ROW_H_SECTION
r += 1

ws.merge_cells(f"A{r}:H{r}")
write(r, 1,
      "All 12 tasks (MFT-1670, MFT-1672, MFT-1684, MFT-1692, MFT-1734, MFT-1746, MFT-1747, "
      "MFT-1785, MFT-1786, MFT-1787, MFT-1788, MFT-1806) were added to the Epic Breakdown sheet. "
      "No further action required. Verdicts and status reflect Jira as source of truth.",
      font=body_font(color="1E5C2E"),
      fill=section_fill(GREEN_FILL),
      align=Alignment(horizontal="left", vertical="top", wrap_text=True))
ws.row_dimensions[r].height = 40
r += 2

# ==== SECTION 2 — Updated ====
ws.merge_cells(f"A{r}:H{r}")
write(r, 1, "⚠️  SECTION 2 — In Excel but NOT in Jira  (updated 2026-03-01  |  6 resolved · 2 still missing)",
      font=Font(name="Calibri", bold=True, size=11, color="FFFFFF"),
      fill=section_fill("833C00"),
      align=Alignment(horizontal="left", vertical="center"))
ws.row_dimensions[r].height = ROW_H_SECTION
r += 1

# Headers
sec2_headers = ["Jira Key", "Epic Group", "Status (2026-03-01)", "Assignee (Jira)", "Resolution", "Recommended Action", None, None]
for c, h in enumerate(sec2_headers, 1):
    if h:
        write(r, c, h, font=header_font(), fill=section_fill("833C00"),
              align=Alignment(horizontal="center", vertical="center"))
ws.row_dimensions[r].height = ROW_H_HDR
r += 1

sec2_data = [
    # key, epic_group, status, assignee, resolution, action, fill
    ("MFT-1695", "Direct Debit",        "To DO",      "Unassigned",           "✅ Ticket exists in Jira",         "Assign owner and confirm scope.",                           GREY_FILL),
    ("MFT-1699", "Direct Debit",        "Not found",  "—",                    "⚠️ Still not in Jira",             "HIGH: Verify if deleted/merged. Remove from Excel if gone.",  RED_FILL),
    ("MFT-1700", "Direct Debit",        "Review",     "Vladislav Nikolaev",   "✅ Ticket exists in Jira",         "No action needed — ticket active and progressing.",          GREY_FILL),
    ("MFT-1704", "Offer Management",    "Developing", "Szymon Giermakowski",  "✅ Ticket exists in Jira",         "No action needed — actively being developed.",              GREY_FILL),
    ("MFT-1705", "Offer Management",    "Not found",  "—",                    "⚠️ Still not in Jira",             "Verify if deleted/merged. Remove from Excel if confirmed gone.", RED_FILL),
    ("MFT-1755", "Monitoring & Alerting","Backlog",   "Stanislav Kaychenkov", "✅ Ticket exists in Jira",         "Ticket in Backlog. Prioritise for next sprint if critical.",  GREY_FILL),
    ("MFT-1830", "Cleanup & Migration", "To DO",      "Andrei Dubskii",       "✅ Ticket exists in Jira",         "Ticket created and assigned. No further action.",            GREY_FILL),
    ("MFT-1831", "Cleanup & Migration", "To DO",      "Andrei Dubskii",       "✅ Ticket exists in Jira",         "Ticket created and assigned. No further action.",            GREY_FILL),
]

for key, epic, status, assignee, resolution, action, fill_hex in sec2_data:
    fill = section_fill(fill_hex)
    vals = [key, epic, status, assignee, resolution, action]
    for c, v in enumerate(vals, 1):
        write(r, c, v, font=body_font(), fill=fill, align=wrap_align(), border=thin_border())
    ws.row_dimensions[r].height = ROW_H_DATA
    r += 1
r += 1

# ==== SECTION 3 — RESOLVED ====
ws.merge_cells(f"A{r}:H{r}")
write(r, 1, "✅  SECTION 3 — Status / Assignee Drift  →  RESOLVED (8 updates applied to Epic Breakdown on 2026-02-28)",
      font=Font(name="Calibri", bold=True, size=11, color="FFFFFF"),
      fill=section_fill("375623"),
      align=Alignment(horizontal="left", vertical="center", wrap_text=True))
ws.row_dimensions[r].height = ROW_H_SECTION
r += 1

ws.merge_cells(f"A{r}:H{r}")
write(r, 1,
      "Status + assignee drift resolved for: MFT-1756, MFT-1757, MFT-1758, MFT-1759 (Backlog / Stanislav Kaychenkov), "
      "MFT-1818 (Cancelled → ✅ DONE), MFT-1819 (Backlog / Unassigned), "
      "MFT-1876 (Review / Evgenii Tumanovskii), MFT-1877 (To DO / Unassigned). "
      "All changes applied in Epic Breakdown sheet.",
      font=body_font(color="1E5C2E"),
      fill=section_fill(GREEN_FILL),
      align=Alignment(horizontal="left", vertical="top", wrap_text=True))
ws.row_dimensions[r].height = 44
r += 2

# ==== SECTION 4 — Updated conflicts ====
ws.merge_cells(f"A{r}:H{r}")
write(r, 1, "⚡  SECTION 4 — Verdict vs Jira Status Conflicts  (updated 2026-03-01  |  3 items require action)",
      font=Font(name="Calibri", bold=True, size=11, color="FFFFFF"),
      fill=section_fill(BLUE_MID),
      align=Alignment(horizontal="left", vertical="center"))
ws.row_dimensions[r].height = ROW_H_SECTION
r += 1

sec4_headers = ["Jira Key", "Summary", "Jira Status\n(2026-03-01)", "Assignee", "Conflict Description", "Risk", "Recommended Resolution", "Owner"]
for c, h in enumerate(sec4_headers, 1):
    write(r, c, h, font=header_font(), fill=section_fill(BLUE_MID),
          align=Alignment(horizontal="center", vertical="center", wrap_text=True))
ws.row_dimensions[r].height = ROW_H_HDR
r += 1

sec4_data = [
    (
        "MFT-1779",
        "Fix the rest of discrepancies found by checker",
        "Ready for release",
        "Evgenii Tumanovskii",
        "Jira status is 'Ready for release'. "
        "Excel verdict was 🟠 NEEDS CLARITY — this is now stale. "
        "Jira is source of truth: work is complete and pending release.",
        "🟡 MEDIUM — stale verdict only; no active risk",
        "Update Excel verdict to ✅ DONE (or 🟢 READY pending deploy). "
        "Confirm release date with Evgenii.",
        "PM",
        AMBER_FILL,
    ),
    (
        "MFT-1718",
        "Resolve conflicts of automated and manual DD repayments in Capital",
        "To DO",
        "Szymon Giermakowski",
        "Jira status is 'To DO' (assigned to Szymon). "
        "Excel verdict was 🟡 IN PROGRESS — Jira not transitioned. "
        "Either work has not started, or Jira is behind.",
        "🟡 MEDIUM — if work active, Jira must be updated",
        "Szymon to confirm: if work started, move Jira to 'Developing'. "
        "If not started, revert verdict to 🟠 NEEDS CLARITY.",
        "Szymon Giermakowski",
        AMBER_FILL,
    ),
    (
        "MFT-1816",
        "[QA] Test Case Design & Coverage Definition - Billing SoT",
        "Developing",
        "Lusine Karapetyan",
        "Jira status is 'Developing' (active). "
        "Excel verdict was 🔴 BLOCKED — blocker appears resolved since Jira shows active development.",
        "🟡 MEDIUM — blocker status may be outdated",
        "Confirm with Lusine that blocker is resolved. "
        "If yes, update Excel verdict to 🟡 IN PROGRESS.",
        "Lusine Karapetyan",
        AMBER_FILL,
    ),
]

for key, summary, jira_status, assignee, conflict, risk, resolution, owner, fill_hex in sec4_data:
    fill = section_fill(fill_hex)
    vals = [key, summary, jira_status, assignee, conflict, risk, resolution, owner]
    for c, v in enumerate(vals, 1):
        write(r, c, v, font=body_font(), fill=fill, align=wrap_align(), border=thin_border())
    ws.row_dimensions[r].height = 55
    r += 1

wb.save(XLSX)
print(f"Inconsistencies sheet updated. Saved: {XLSX}")
