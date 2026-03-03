"""
Reassign Unphased rows to their correct phases, remove Unphased category, re-sort.
  Capital as repayments engine        (MFT-1806/07/08/10) → Phase 2
  Capital should fully rely on billing (MFT-1676/17)       → Phase 2
  Tech debt (make capital great again) (MFT-1692)          → Phase 3
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

FILE  = "outputs/tech_debt_report.xlsx"
SHEET = "Epic Breakdown"
DATA_START = 4
JIRA_BASE  = "https://tabby.atlassian.net/browse/"

# ── Phase colours (no Unphased any more) ────────────────────────────────────
PHASE_FILL = {
    "Phase 0":  "FF2C3E50",
    "Phase 1":  "FF1A5276",
    "Phase 2":  "FF1E8449",
    "Phase 3":  "FF6C3483",
    "Parallel": "FF117A65",
}
PHASE_ORDER = {"Phase 0": 0, "Phase 1": 1, "Phase 2": 2,
               "Phase 3": 3, "Parallel": 4}

# ── Unphased epic → correct phase ───────────────────────────────────────────
REASSIGN = {
    "Capital as repayments engine":        "Phase 2",
    "Capital should fully rely on billing data": "Phase 2",
    "Tech debt (make capital great again)": "Phase 3",
}

# ── Quality colours (col I) ──────────────────────────────────────────────────
QUALITY_FILL = {
    "✓ Complete":          ("FF375623", "FFFFFFFF"),
    "OK":                  ("FF375623", "FFFFFFFF"),
    "Partial (1 missing)": ("FF7F6000", "FFFFFFFF"),
    "Partial (2 missing)": ("FF833C00", "FFFFFFFF"),
    "✗ All missing":       ("FF9C0006", "FFFFFFFF"),
}

BAND_ODD  = "FFFAFAFA"
BAND_EVEN = "FFFFFFFF"

def make_fill(hex8):
    return PatternFill(fill_type="solid", fgColor=hex8)

def thin_border():
    s = Side(style="thin", color="FFD9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)

def phase_key(p):   return PHASE_ORDER.get(p, 99)
def epic_key(e):    return (e or "").lower()
def issue_num(k):
    if k and "-" in k:
        try: return int(k.split("-")[1])
        except ValueError: pass
    return 0

# ── Load ─────────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(FILE)
ws = wb[SHEET]

last_row = DATA_START - 1
for r in range(DATA_START, 1000):
    if ws.cell(r, 3).value is None:
        break
    last_row = r
print(f"  Data rows: {DATA_START}–{last_row} ({last_row - DATA_START + 1} rows)")

# ── Read all rows, fix Unphased ──────────────────────────────────────────────
rows = []
reassigned = []

for r in range(DATA_START, last_row + 1):
    phase = ws.cell(r, 3).value or ""
    epic  = ws.cell(r, 1).value or ""

    if phase == "Unphased":
        new_phase = REASSIGN.get(epic, "Phase 3")  # safe fallback
        reassigned.append((ws.cell(r, 4).value, epic, phase, new_phase))
        phase = new_phase

    rows.append({
        "phase": phase,
        "epic":  epic,
        "cols":  [ws.cell(r, c).value for c in range(1, 14)],
    })

# ── Sort: Phase 0 → 1 → 2 → 3 → Parallel ────────────────────────────────────
rows.sort(key=lambda rd: (phase_key(rd["phase"]),
                          epic_key(rd["epic"]),
                          issue_num(rd["cols"][3])))

# ── Write back ───────────────────────────────────────────────────────────────
for i, rd in enumerate(rows):
    r         = DATA_START + i
    phase     = rd["phase"]
    vals      = rd["cols"]
    is_odd    = (i % 2 == 0)
    phase_hex = PHASE_FILL.get(phase, "FF2C3E50")
    band_hex  = BAND_ODD if is_odd else BAND_EVEN

    for c in range(1, 14):
        cell = ws.cell(r, c)
        val  = vals[c - 1]

        if c == 1:  # Phase chip
            cell.value = val
            cell.fill  = make_fill(phase_hex)
            cell.font  = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border()

        elif c == 2:  # Epic Link
            cell.value = val
            if val and val != "—" and str(val).startswith("MFT-"):
                cell.hyperlink = JIRA_BASE + val
                cell.font = Font(name="Calibri", size=11, color="FF0563C1", underline="single")
            else:
                cell.font = Font(name="Calibri", size=11, color="FF000000")
            cell.fill      = make_fill(band_hex)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = thin_border()

        elif c == 3:  # Phase label
            cell.value = phase
            cell.fill  = make_fill(phase_hex)
            cell.font  = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border()

        elif c == 4:  # Issue Key
            cell.value = val
            if val and str(val).startswith("MFT-"):
                cell.hyperlink = JIRA_BASE + val
                cell.font = Font(name="Calibri", size=11, color="FF0563C1", underline="single")
            else:
                cell.font = Font(name="Calibri", size=11, color="FF000000")
            cell.fill      = make_fill(band_hex)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = thin_border()

        elif c == 9:  # Desc Quality
            cell.value = val
            fill_hex, text_hex = QUALITY_FILL.get(
                str(val).strip() if val else "", ("FF7F6000", "FFFFFFFF"))
            cell.fill  = make_fill(fill_hex)
            cell.font  = Font(name="Calibri", size=10, bold=True, color=text_hex)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border()

        else:  # plain banding
            cell.value = val
            cell.fill  = make_fill(band_hex)
            cell.font  = Font(name="Calibri", size=11, color="FF000000")
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = thin_border()

wb.save(FILE)

print(f"\n  Saved: {FILE}")
print(f"\n  Reassigned rows:")
for key, epic, old, new in reassigned:
    print(f"    {key:<12} {old} → {new}  ({epic})")

# ── Verify ───────────────────────────────────────────────────────────────────
from collections import Counter
wb2  = openpyxl.load_workbook(FILE)
ws2  = wb2[SHEET]
dist = Counter()
for r in range(DATA_START, 1000):
    p = ws2.cell(r, 3).value
    if p is None: break
    dist[p] += 1

print(f"\n  Final phase distribution:")
for p, n in sorted(dist.items(), key=lambda x: PHASE_ORDER.get(x[0], 99)):
    print(f"    {p:<12}: {n} rows")
print(f"    {'TOTAL':<12}: {sum(dist.values())} rows")
print(f"\n  'Unphased' remaining: {dist.get('Unphased', 0)}")
