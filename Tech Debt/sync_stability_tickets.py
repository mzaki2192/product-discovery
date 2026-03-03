#!/usr/bin/env python3
"""Sync Jira stability-labeled tickets into the Tech Debt report."""
import base64
import json
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

REPORT_PATH = Path('/Users/zaki/Downloads/Product discovery/Tech Debt/outputs/tech_debt_report.xlsx')
SHEET = 'Epic Breakdown'
DATA_START_ROW = 4


def load_jira_config():
    cfg_path = Path('/Users/zaki/.claude.json')
    cfg = json.loads(cfg_path.read_text(encoding='utf-8'))
    jira_cfg = cfg['mcpServers']['jira']['env']
    return jira_cfg['JIRA_URL'].rstrip('/'), jira_cfg['JIRA_USERNAME'], jira_cfg['JIRA_API_TOKEN']


def jira_search_all_stability(jira_url, user, token):
    auth_header = 'Basic ' + base64.b64encode(f'{user}:{token}'.encode()).decode()
    base_url = f"{jira_url}/rest/api/3/search"
    jql = 'labels = stability ORDER BY key ASC'
    fields = 'summary,status,assignee,labels,issuetype,parent,description'

    start_at = 0
    max_results = 100
    all_issues = []

    while True:
        params = {
            'jql': jql,
            'fields': fields,
            'startAt': start_at,
            'maxResults': max_results,
        }
        url = base_url + '?' + urllib.parse.urlencode(params)
        req = urllib.request.Request(
            url,
            headers={
                'Authorization': auth_header,
                'Accept': 'application/json',
            },
            method='GET',
        )
        with urllib.request.urlopen(req, timeout=60) as resp:
            data = json.loads(resp.read())

        issues = data.get('issues', [])
        all_issues.extend(issues)

        total = int(data.get('total', 0))
        start_at += len(issues)
        if not issues or start_at >= total:
            break

    return all_issues


def assignee_name(issue):
    fields = issue.get('fields', {})
    a = fields.get('assignee')
    if not a:
        return '—'
    return a.get('displayName') or a.get('display_name') or a.get('name') or '—'


def issue_type(issue):
    return (issue.get('fields', {}).get('issuetype') or {}).get('name') or 'Task'


def issue_status(issue):
    return (issue.get('fields', {}).get('status') or {}).get('name') or 'Unknown'


def issue_summary(issue):
    return issue.get('fields', {}).get('summary') or ''


def parent_key(issue):
    p = issue.get('fields', {}).get('parent')
    if not p:
        return '—'
    return p.get('key') or '—'


def parent_summary(issue):
    p = issue.get('fields', {}).get('parent')
    if not p:
        return issue_summary(issue)
    pf = p.get('fields', {})
    return pf.get('summary') or parent_key(issue)


def find_row_by_key(ws):
    row_by_key = {}
    for r in range(DATA_START_ROW, ws.max_row + 1):
        key = ws.cell(r, 4).value  # Col D Issue Key
        if isinstance(key, str) and key.strip().startswith('MFT-'):
            row_by_key[key.strip()] = r
    return row_by_key


def first_empty_row(ws):
    r = ws.max_row + 1
    while any(ws.cell(r, c).value not in (None, '') for c in range(1, 13)):
        r += 1
    return r


def main():
    jira_url, user, token = load_jira_config()
    issues = jira_search_all_stability(jira_url, user, token)
    issues_by_key = {i['key']: i for i in issues}

    wb = load_workbook(REPORT_PATH)
    ws = wb[SHEET]

    row_by_key = find_row_by_key(ws)
    report_keys = set(row_by_key.keys())
    jira_keys = set(issues_by_key.keys())

    missing_in_report = sorted(jira_keys - report_keys)
    not_in_jira_anymore = sorted(report_keys - jira_keys)

    updated_rows = 0
    for key in sorted(jira_keys & report_keys):
        row = row_by_key[key]
        issue = issues_by_key[key]

        changes = 0
        # E=issue type F=summary G=status H=assignee
        val_map = {
            5: issue_type(issue),
            6: issue_summary(issue),
            7: issue_status(issue),
            8: assignee_name(issue),
        }
        for col, new_val in val_map.items():
            if ws.cell(row, col).value != new_val:
                ws.cell(row, col).value = new_val
                changes += 1
        if changes:
            updated_rows += 1

    appended = 0
    for key in missing_in_report:
        issue = issues_by_key[key]
        r = first_empty_row(ws)
        # A Epic, B Epic Link, C Phase, D Key, E type, F summary, G status, H assignee
        ws.cell(r, 1).value = parent_summary(issue)
        ws.cell(r, 2).value = parent_key(issue)
        ws.cell(r, 3).value = 'Unmapped'
        ws.cell(r, 4).value = key
        ws.cell(r, 5).value = issue_type(issue)
        ws.cell(r, 6).value = issue_summary(issue)
        ws.cell(r, 7).value = issue_status(issue)
        ws.cell(r, 8).value = assignee_name(issue)
        ws.cell(r, 9).value = None
        ws.cell(r, 10).value = '🟠 NEEDS CLARITY'
        ws.cell(r, 11).value = None
        ws.cell(r, 12).value = 'Auto-added from Jira stability label sync on ' + datetime.now().strftime('%Y-%m-%d')
        appended += 1

    ws.cell(2, 1).value = (
        f"Stability sync complete | Jira stability tickets: {len(jira_keys)} | "
        f"Updated: {updated_rows} | Added: {appended} | Generated: {datetime.now().strftime('%Y-%m-%d')}"
    )

    wb.save(REPORT_PATH)

    print(f'Stability issues from Jira: {len(jira_keys)}')
    print(f'Existing in report: {len(report_keys)}')
    print(f'Updated rows: {updated_rows}')
    print(f'Added rows: {appended}')
    print(f'In report but not in Jira stability now: {len(not_in_jira_anymore)}')
    if missing_in_report:
        print('Added keys:', ', '.join(missing_in_report))
    if not_in_jira_anymore:
        print('Report keys not currently in stability label set:', ', '.join(not_in_jira_anymore[:25]))


if __name__ == '__main__':
    main()
