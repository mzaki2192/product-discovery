"""
Comprehensive fix for Epic Breakdown sheet:
1. Remove duplicate MFT-1805 (the one with no/broken fill)
2. Sort rows: Phase 0 → Phase 1 → Phase 2 → Phase 3 → Various, then by Epic, then by Issue Key number
3. Fix col D font: blue (#0563C1) + single underline for all rows
4. Fix col E fill: apply alternating row banding (FFFAFAFA / FFFFFFFF) matching all other data cols
5. Fix col A fill: correct phase-appropriate fill with opaque alpha (FF prefix)
6. Re-apply alternating banding on cols B-M for all rows after sort
"""
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX = "/Users/zaki/Downloads/Product discovery/Tech Debt/outputs/tech_debt_report.xlsx"
JIRA_BASE = "https://tabby.atlassian.net/browse/"

wb = openpyxl.load_workbook(XLSX)
ws = wb["Epic Breakdown"]

# ---- Phase ordering ----
PHASE_ORDER = {"Phase 0": 0, "Phase 1": 1, "Phase 2": 2, "Phase 3": 3, "Various": 4}

def phase_sort_key(phase_val):
    if not phase_val:
        return 99
    p = str(phase_val).strip()
    return PHASE_ORDER.get(p, 99)

def issue_key_num(key_val):
    """Extract numeric part of MFT-XXXX for sorting."""
    if not key_val:
        return 99999
    m = re.search(r'(\d+)', str(key_val))
    return int(m.group(1)) if m else 99999

# ---- Phase → col A fill color (opaque) ----
PHASE_FILL = {
    "Phase 0": "FF2C3E50",   # dark navy
    "Phase 1": "FF1A5276",   # dark blue
    "Phase 2": "FF1E8449",   # green
    "Phase 3": "FF6C3483",   # purple
    "Various": "FF2C3E50",   # dark navy (same as Phase 0)
}
DEFAULT_COL_A_FILL = "FF2C3E50"

# Alternating banding for data cols B-M
BAND_ODD  = "FFFAFAFA"
BAND_EVEN = "FFFFFFFF"

# ---- Read header rows (rows 1-3) — keep as-is ----
# Detect actual data start (row 4)
DATA_START = 4
DATA_END   = ws.max_row

MAX_COL = ws.max_column  # should be 13 (A-M)

print(f"Sheet dimensions: rows 1-{DATA_END}, cols 1-{MAX_COL}")

# ---- Collect all data rows ----
def read_row(row_idx):
    """Return dict with all cell data for a given row."""
    cells = []
    for col in range(1, MAX_COL + 1):
        cell = ws.cell(row_idx, col)
        cells.append({
            "value": cell.value,
            "font": cell.font.copy(),
            "fill": cell.fill.copy() if cell.fill else None,
            "alignment": cell.alignment.copy(),
            "border": cell.border.copy(),
            "hyperlink": cell.hyperlink.target if cell.hyperlink else None,
            "number_format": cell.number_format,
        })
    return cells

rows_data = []
for row_idx in range(DATA_START, DATA_END + 1):
    row = read_row(row_idx)
    # Col C (index 2) = Phase, Col D (index 3) = Issue Key
    phase_val = row[2]["value"]
    key_val   = row[3]["value"]
    rows_data.append({
        "row_idx": row_idx,
        "phase": str(phase_val).strip() if phase_val else "",
        "key": str(key_val).strip() if key_val else "",
        "epic": str(row[0]["value"]).strip() if row[0]["value"] else "",
        "cells": row,
    })

print(f"Read {len(rows_data)} data rows")

# ---- Identify and remove duplicate MFT-1805 ----
# Keep the one with a real fill (col A has opaque fill), remove the broken one
mft1805_rows = [r for r in rows_data if r["key"] == "MFT-1805"]
print(f"MFT-1805 occurrences: {len(mft1805_rows)} (at original rows {[r['row_idx'] for r in mft1805_rows]})")

if len(mft1805_rows) > 1:
    def fill_is_valid(row_data):
        """Check if col A fill has opaque alpha."""
        fill = row_data["cells"][0]["fill"]
        if fill is None:
            return False
        fg = fill.fgColor
        if fg is None:
            return False
        # Check if it's a solid fill with opaque color
        if fill.fill_type not in (None, "none"):
            rgb = str(fg.rgb) if fg.type == "rgb" else ""
            return rgb.startswith("FF") and len(rgb) == 8
        return False

    valid = [r for r in mft1805_rows if fill_is_valid(r)]
    invalid = [r for r in mft1805_rows if not fill_is_valid(r)]

    if invalid:
        for bad in invalid:
            rows_data = [r for r in rows_data if not (r["key"] == "MFT-1805" and r["row_idx"] == bad["row_idx"])]
            print(f"  Removed duplicate MFT-1805 at original row {bad['row_idx']}")
    else:
        # Remove the one appearing later (the re-appended one)
        last = max(mft1805_rows, key=lambda r: r["row_idx"])
        rows_data = [r for r in rows_data if not (r["key"] == "MFT-1805" and r["row_idx"] == last["row_idx"])]
        print(f"  Removed duplicate MFT-1805 at original row {last['row_idx']} (both valid, kept earlier)")

print(f"Rows after dedup: {len(rows_data)}")

# ---- Sort rows ----
def sort_key(r):
    phase = r["phase"]
    epic  = r["epic"].lower()
    num   = issue_key_num(r["key"])
    return (phase_sort_key(phase), epic, num)

rows_data.sort(key=sort_key)

# Verify sort
prev_phase_order = -1
for r in rows_data:
    po = phase_sort_key(r["phase"])
    if po < prev_phase_order:
        print(f"  WARNING: sort issue at key={r['key']} phase={r['phase']}")
    prev_phase_order = po

print("Sort complete.")

# ---- Clear all data rows in sheet ----
for row_idx in range(DATA_START, DATA_END + 1):
    for col in range(1, MAX_COL + 1):
        cell = ws.cell(row_idx, col)
        cell.value = None
        cell.font = Font()
        cell.fill = PatternFill()
        cell.alignment = Alignment()
        cell.border = Border()
        cell.hyperlink = None

# ---- Style helpers ----
def thin_border():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def band_fill(row_position):
    """row_position is 1-based relative to data start."""
    hex_color = BAND_ODD if row_position % 2 == 1 else BAND_EVEN
    return PatternFill("solid", fgColor=hex_color)

def col_a_fill(phase):
    hex_color = PHASE_FILL.get(phase, DEFAULT_COL_A_FILL)
    return PatternFill("solid", fgColor=hex_color)

def hyperlink_font():
    return Font(name="Calibri", size=11, color="FF0563C1", underline="single")

# ---- Write sorted rows back ----
for pos, r in enumerate(rows_data, 1):
    write_row = DATA_START + pos - 1
    cells = r["cells"]
    phase = r["phase"]
    key   = r["key"]

    for col_idx in range(1, MAX_COL + 1):
        cell_data = cells[col_idx - 1]
        out_cell = ws.cell(write_row, col_idx)
        out_cell.value = cell_data["value"]
        out_cell.number_format = cell_data["number_format"]

        # Col A: phase-colored fill, white bold font
        if col_idx == 1:
            out_cell.fill = col_a_fill(phase)
            out_cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            out_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            out_cell.border = thin_border()

        # Col D: Issue Key — always hyperlink blue underline
        elif col_idx == 4:
            if key:
                out_cell.hyperlink = JIRA_BASE + key
            out_cell.font = hyperlink_font()
            out_cell.fill = band_fill(pos)
            out_cell.alignment = Alignment(horizontal="center", vertical="center")
            out_cell.border = thin_border()

        # All other data cols (B, C, E-M): alternating banding
        else:
            out_cell.fill = band_fill(pos)
            # Preserve original font but ensure it's not transparent/broken
            orig_font = cell_data["font"]
            # For safety, preserve original font settings but reset any bad color
            out_cell.font = Font(
                name=orig_font.name or "Calibri",
                size=orig_font.size or 11,
                bold=orig_font.bold,
                italic=orig_font.italic,
                color=orig_font.color.rgb if (orig_font.color and orig_font.color.type == "rgb") else "000000",
            )
            out_cell.alignment = Alignment(
                horizontal=cell_data["alignment"].horizontal or "left",
                vertical=cell_data["alignment"].vertical or "top",
                wrap_text=cell_data["alignment"].wrap_text,
            )
            out_cell.border = thin_border()

    ws.row_dimensions[write_row].height = 30

wb.save(XLSX)
print(f"\nDone. Saved: {XLSX}")
print(f"  Total data rows written: {len(rows_data)}")

# Quick verification
wb2 = openpyxl.load_workbook(XLSX)
ws2 = wb2["Epic Breakdown"]
phases_seen = []
for row_idx in range(DATA_START, DATA_START + len(rows_data)):
    phase_cell = ws2.cell(row_idx, 3)
    phases_seen.append((row_idx, phase_cell.value))

# Print phase transitions
prev = None
for row_idx, pv in phases_seen:
    if pv != prev:
        print(f"  Phase transition at row {row_idx}: {pv}")
        prev = pv
