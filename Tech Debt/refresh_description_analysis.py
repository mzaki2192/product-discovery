"""
Refresh description analysis in the Tech Debt report from live Jira data.

Updates in sheet "Epic Breakdown":
- Col I  (Desc Quality)
- Col K  (Description Analysis)
- Row 2 generated date text
"""
import base64
import json
import re
import urllib.request
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


REPORT_PATH = "/Users/zaki/Downloads/Product discovery/Tech Debt/outputs/tech_debt_report.xlsx"
SHEET = "Epic Breakdown"
DATA_START_ROW = 4

# Jira config reused from local MCP config file.
with open("/Users/zaki/.claude.json", "r", encoding="utf-8") as f:
    cfg = json.load(f)
jira_cfg = cfg["mcpServers"]["jira"]["env"]
JIRA_URL = jira_cfg["JIRA_URL"].rstrip("/")
JIRA_USER = jira_cfg["JIRA_USERNAME"]
JIRA_TOKEN = jira_cfg["JIRA_API_TOKEN"]
JIRA_SEARCH_URL = f"{JIRA_URL}/rest/api/3/search/jql"
AUTH_HEADER = "Basic " + base64.b64encode(f"{JIRA_USER}:{JIRA_TOKEN}".encode()).decode()


QUALITY_STYLES = {
    "✓ Complete": ("FF375623", "FFFFFFFF"),
    "OK": ("FF375623", "FFFFFFFF"),
    "Partial (1 missing)": ("FF7F6000", "FFFFFFFF"),
    "Partial (2 missing)": ("FF833C00", "FFFFFFFF"),
    "✗ All missing": ("FF9C0006", "FFFFFFFF"),
}
DEFAULT_STYLE = ("FF7F6000", "FFFFFFFF")


def jira_search(jql, fields, max_results=50):
    payload = json.dumps(
        {
            "jql": jql,
            "fields": fields,
            "maxResults": max_results,
        }
    ).encode()
    req = urllib.request.Request(
        JIRA_SEARCH_URL,
        data=payload,
        headers={
            "Authorization": AUTH_HEADER,
            "Accept": "application/json",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=45) as resp:
        return json.loads(resp.read())


def normalize_description(desc):
    if desc is None:
        return ""
    if isinstance(desc, str):
        return desc
    # Jira Cloud can return ADF objects. Keep a compact, deterministic text fallback.
    return json.dumps(desc, ensure_ascii=False)


def has_context(text):
    return bool(re.search(r"\b(context|problem|why)\b", text, flags=re.IGNORECASE))


def has_outcome(text):
    return bool(
        re.search(
            r"\b(outcome|definition of done|dod|acceptance criteria|exit criteria)\b",
            text,
            flags=re.IGNORECASE,
        )
    )


def has_approach(text):
    return bool(
        re.search(
            r"\b(approach|technical approach|implementation|scope)\b",
            text,
            flags=re.IGNORECASE,
        )
    )


def assess_description(text):
    text = text or ""
    c = has_context(text)
    o = has_outcome(text)
    a = has_approach(text)

    missing = sum([not c, not o, not a])
    if missing == 0:
        quality = "✓ Complete"
    elif missing == 1:
        quality = "Partial (1 missing)"
    elif missing == 2:
        quality = "Partial (2 missing)"
    else:
        quality = "✗ All missing"

    analysis = (
        "Template Assessment:\n"
        f"  Context:     {'✓ Present' if c else '✗ Missing'}\n"
        f"  Outcome/DoD: {'✓ Present' if o else '✗ Missing'}\n"
        f"  Approach:    {'✓ Present' if a else '✗ Missing'}\n"
        f"  Desc length: {len(text)} chars"
    )
    return quality, analysis


def style_quality_cell(cell, quality):
    fill_hex, text_hex = QUALITY_STYLES.get(quality, DEFAULT_STYLE)
    cell.fill = PatternFill(fill_type="solid", fgColor=fill_hex)
    cell.font = Font(name="Calibri", size=10, bold=True, color=text_hex)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    side = Side(style="thin", color="FFD9D9D9")
    cell.border = Border(left=side, right=side, top=side, bottom=side)


def main():
    wb = load_workbook(REPORT_PATH)
    ws = wb[SHEET]

    # Collect keys from report (col D).
    keys = []
    row_by_key = {}
    for r in range(DATA_START_ROW, ws.max_row + 1):
        key = ws.cell(r, 4).value
        if isinstance(key, str) and key.startswith("MFT-"):
            key = key.strip()
            keys.append(key)
            row_by_key[key] = r

    uniq_keys = []
    seen = set()
    for k in keys:
        if k not in seen:
            seen.add(k)
            uniq_keys.append(k)

    # Fetch latest Jira descriptions in batches.
    latest_desc = {}
    batch_size = 50
    for i in range(0, len(uniq_keys), batch_size):
        batch = uniq_keys[i : i + batch_size]
        jql = "issue in ({})".format(", ".join(batch))
        data = jira_search(jql, fields=["description"], max_results=batch_size)
        for issue in data.get("issues", []):
            key = issue["key"]
            desc = issue.get("fields", {}).get("description")
            latest_desc[key] = normalize_description(desc)

    # Write quality + analysis into report.
    counts = {
        "✓ Complete": 0,
        "Partial (1 missing)": 0,
        "Partial (2 missing)": 0,
        "✗ All missing": 0,
    }
    missing_from_jira = []

    for key in uniq_keys:
        row = row_by_key[key]
        if key not in latest_desc:
            missing_from_jira.append(key)
            desc_text = ""
        else:
            desc_text = latest_desc[key]

        quality, analysis = assess_description(desc_text)
        ws.cell(row, 9).value = quality
        ws.cell(row, 11).value = analysis
        style_quality_cell(ws.cell(row, 9), quality)
        counts[quality] += 1

    # Update generated date line in row 2 (col A).
    now_str = datetime.now().strftime("%Y-%m-%d")
    ws.cell(
        2, 1
    ).value = (
        "All 108 tasks enriched with GitLab + NotebookLM + Granola + BigQuery context"
        f"   |   Generated: {now_str}"
    )

    wb.save(REPORT_PATH)

    print(f"Updated workbook: {REPORT_PATH}")
    print(f"Issues processed: {len(uniq_keys)}")
    print("Quality counts:")
    for k, v in counts.items():
        print(f"  {k}: {v}")
    if missing_from_jira:
        print(f"Missing from Jira response ({len(missing_from_jira)}): {', '.join(missing_from_jira)}")


if __name__ == "__main__":
    main()
