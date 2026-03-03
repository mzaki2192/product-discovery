"""
Replace "Various" phase with "Parallel" or "Unphased" in Epic Breakdown sheet.
- Updates col C (Phase text)
- Updates col A and col C fill colors
- Re-sorts data: Phase 0 → 1 → 2 → 3 → Parallel → Unphased
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from copy import copy

FILE = "outputs/tech_debt_report.xlsx"
SHEET = "Epic Breakdown"
DATA_START = 4
JIRA_BASE = "https://tabby.atlassian.net/browse/"

# ── Phase fill colours ──────────────────────────────────────────────────────
PHASE_FILL = {
    "Phase 0":  "FF2C3E50",  # dark navy
    "Phase 1":  "FF1A5276",  # dark blue
    "Phase 2":  "FF1E8449",  # dark green
    "Phase 3":  "FF6C3483",  # dark purple
    "Parallel": "FF117A65",  # dark teal   ← intentional parallel workstream
    "Unphased": "FF784212",  # dark burnt orange ← needs phase assignment
}
PHASE_ORDER = {"Phase 0": 0, "Phase 1": 1, "Phase 2": 2,
               "Phase 3": 3, "Parallel": 4, "Unphased": 5}

# ── Which "Various" epics map to which new phase ────────────────────────────
PARALLEL_EPICS = {
    "[QA] QA Planning & Release Readiness for Billing Migration",
    "Build Capital->Bank statement reconciliation",
    "Build Capital->Reem reconciliation",
    "Build first iteration of test model / Lifecycle Integrity Checks for loans",
    "Fix all discrepancies with billing",
    "Increate Capital Observability",
    "Increate observability of application",
    "Update Capital->Billing reconciliation",
}
UNPHASED_EPICS = {
    "Capital as repayments engine",
    "Capital should fully rely on billing data",
    "Tech debt (make capital great again)",
}

# ── Quality fill (col I) ────────────────────────────────────────────────────
QUALITY_FILL = {
    "✓ Complete":          ("FF375623", "FFFFFFFF"),
    "OK":                  ("FF375623", "FFFFFFFF"),
    "Partial (1 missing)": ("FF7F6000", "FFFFFFFF"),
    "Partial (2 missing)": ("FF833C00", "FFFFFFFF"),
    "✗ All missing":       ("FF9C0006", "FFFFFFFF"),
}

# ── Banding colours (non-special cols) ─────────────────────────────────────
BAND_ODD  = "FFFAFAFA"
BAND_EVEN = "FFFFFFFF"

# ── Helpers ─────────────────────────────────────────────────────────────────
def make_fill(hex8):
    return PatternFill(fill_type="solid", fgColor=hex8)

def thin_border():
    s = Side(style="thin", color="FFD9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)

def phase_key(p):
    return PHASE_ORDER.get(p, 99)

def epic_key(e):
    return (e or "").lower()

def issue_num(k):
    if k and "-" in k:
        try:
            return int(k.split("-")[1])
        except ValueError:
            pass
    return 0

# ── Load workbook ────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(FILE)
ws = wb[SHEET]

# ── Find last data row ───────────────────────────────────────────────────────
last_row = DATA_START - 1
for r in range(DATA_START, 1000):
    if ws.cell(r, 3).value is None:
        break
    last_row = r

print(f"  Data rows found: {DATA_START}–{last_row} ({last_row - DATA_START + 1} rows)")

# ── Read all rows ────────────────────────────────────────────────────────────
rows = []
for r in range(DATA_START, last_row + 1):
    phase = ws.cell(r, 3).value or ""
    epic  = ws.cell(r, 1).value or ""

    # Reclassify "Various"
    if phase == "Various":
        if epic in PARALLEL_EPICS:
            phase = "Parallel"
        elif epic in UNPHASED_EPICS:
            phase = "Unphased"
        else:
            phase = "Unphased"  # safe fallback

    row_data = {
        "phase": phase,
        "epic":  epic,
        "cols":  [ws.cell(r, c).value for c in range(1, 14)],  # A–M
    }
    rows.append(row_data)

# ── Sort ─────────────────────────────────────────────────────────────────────
rows.sort(key=lambda rd: (phase_key(rd["phase"]), epic_key(rd["epic"]),
                          issue_num(rd["cols"][3])))  # col D = index 3

# ── Write rows back ──────────────────────────────────────────────────────────
reclassified = {"Parallel": 0, "Unphased": 0}

for i, rd in enumerate(rows):
    r       = DATA_START + i
    phase   = rd["phase"]
    vals    = rd["cols"]
    is_odd  = (i % 2 == 0)  # 0-indexed → first row is odd position

    if phase in reclassified:
        reclassified[phase] += 1

    phase_hex = PHASE_FILL.get(phase, "FF2C3E50")
    band_hex  = BAND_ODD if is_odd else BAND_EVEN

    for c in range(1, 14):
        cell = ws.cell(r, c)
        val  = vals[c - 1]

        # ── col A (Phase chip) ───────────────────────────────────────────
        if c == 1:
            cell.value = val
            cell.fill  = make_fill(phase_hex)
            cell.font  = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border = thin_border()

        # ── col B (Epic Link — hyperlink) ────────────────────────────────
        elif c == 2:
            cell.value = val
            if val and val != "—" and val.startswith("MFT-"):
                cell.hyperlink = JIRA_BASE + val
                cell.font = Font(name="Calibri", size=11,
                                 color="FF0563C1", underline="single")
            else:
                cell.font = Font(name="Calibri", size=11, color="FF000000")
            cell.fill      = make_fill(band_hex)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = thin_border()

        # ── col C (Phase label) ──────────────────────────────────────────
        elif c == 3:
            cell.value = phase          # updated value
            cell.fill  = make_fill(phase_hex)
            cell.font  = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border()

        # ── col D (Issue Key — hyperlink) ────────────────────────────────
        elif c == 4:
            cell.value = val
            if val and val.startswith("MFT-"):
                cell.hyperlink = JIRA_BASE + val
                cell.font = Font(name="Calibri", size=11,
                                 color="FF0563C1", underline="single")
            else:
                cell.font = Font(name="Calibri", size=11, color="FF000000")
            cell.fill      = make_fill(band_hex)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = thin_border()

        # ── col I (Desc Quality — colour-coded) ──────────────────────────
        elif c == 9:
            cell.value = val
            fill_hex, text_hex = QUALITY_FILL.get(
                str(val).strip() if val else "",
                ("FF7F6000", "FFFFFFFF")
            )
            cell.fill  = make_fill(fill_hex)
            cell.font  = Font(name="Calibri", size=10, bold=True, color=text_hex)
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border = thin_border()

        # ── all other cols — plain banding ───────────────────────────────
        else:
            cell.value = val
            cell.fill  = make_fill(band_hex)
            cell.font  = Font(name="Calibri", size=11, color="FF000000")
            cell.alignment = Alignment(horizontal="left", vertical="center",
                                       wrap_text=True)
            cell.border = thin_border()

wb.save(FILE)
print(f"\n  Done. Saved: {FILE}")
print(f"  Reclassified → Parallel : {reclassified['Parallel']} rows")
print(f"  Reclassified → Unphased : {reclassified['Unphased']} rows")
print(f"  Sort order   : Phase 0 → 1 → 2 → 3 → Parallel → Unphased")
