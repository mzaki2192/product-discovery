"""
Inserts a new "Issue Type" column (col E) immediately after "Issue Key" (col D)
in the "Epic Breakdown" sheet of tech_debt_report.xlsx.
Reads issue types from /tmp/issue_types.json (produced by fetch_issue_types.py).
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

XLSX = "/Users/zaki/Downloads/Product discovery/Tech Debt/outputs/tech_debt_report.xlsx"
ISSUE_TYPES_JSON = "/tmp/issue_types.json"

with open(ISSUE_TYPES_JSON) as f:
    issue_types = json.load(f)

wb = openpyxl.load_workbook(XLSX)
ws = wb["Epic Breakdown"]

# Verify current layout: col D should be "Issue Key"
# Headers are in rows 1-3; data starts row 4
# Find which row has "Issue Key" header
issue_key_col = None
for row in range(1, 4):
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row, col).value
        if val and "issue key" in str(val).lower():
            issue_key_col = col
            header_row = row
            print(f"Found 'Issue Key' header at row {row}, col {col}")
            break
    if issue_key_col:
        break

if not issue_key_col:
    print("ERROR: Could not find 'Issue Key' column header")
    exit(1)

# Insert new column E (after col D = issue_key_col)
new_col = issue_key_col + 1
ws.insert_cols(new_col)
print(f"Inserted new column at position {new_col}")

# --- Copy header styling from the Issue Key header cell ---
src_header = ws.cell(header_row, issue_key_col)

# Write header "Issue Type" in the new column, matching existing header style
hdr_cell = ws.cell(header_row, new_col)
hdr_cell.value = "Issue Type"

if src_header.font:
    hdr_cell.font = Font(
        name=src_header.font.name,
        size=src_header.font.size,
        bold=src_header.font.bold,
        color=src_header.font.color.rgb if src_header.font.color and src_header.font.color.type == "rgb" else "000000",
    )
if src_header.fill and src_header.fill.fill_type != "none":
    hdr_cell.fill = PatternFill(
        fill_type=src_header.fill.fill_type,
        fgColor=src_header.fill.fgColor,
        bgColor=src_header.fill.bgColor,
    )
if src_header.alignment:
    hdr_cell.alignment = Alignment(
        horizontal=src_header.alignment.horizontal,
        vertical=src_header.alignment.vertical,
        wrap_text=src_header.alignment.wrap_text,
    )

# Also copy header cells in other header rows (rows 1-3) that we didn't land on
for r in range(1, 4):
    if r == header_row:
        continue
    src = ws.cell(r, issue_key_col)
    tgt = ws.cell(r, new_col)
    # Merge info — just leave blank, will inherit sheet style

# --- Populate data rows ---
written = 0
missing_keys = []

for row_idx in range(4, ws.max_row + 1):
    key_cell = ws.cell(row_idx, issue_key_col)
    key_val = key_cell.value
    if not key_val:
        continue

    # Strip any hyperlink text — key_val should already be plain key
    key_str = str(key_val).strip()
    issue_type = issue_types.get(key_str)

    tgt = ws.cell(row_idx, new_col)
    if issue_type:
        tgt.value = issue_type
        written += 1
    else:
        tgt.value = "—"
        missing_keys.append(key_str)

    # Match font/alignment of neighbouring cell
    tgt.font = Font(name="Calibri", size=11)
    tgt.alignment = Alignment(horizontal="left", vertical="center")

# Set column width
ws.column_dimensions[openpyxl.utils.get_column_letter(new_col)].width = 14

wb.save(XLSX)
print(f"\nDone. Written {written} issue types.")
if missing_keys:
    print(f"Missing ({len(missing_keys)}): {missing_keys}")
else:
    print("No missing keys.")
print(f"Saved: {XLSX}")
