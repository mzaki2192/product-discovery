"""
Ensures every Issue Key cell in col D (Epic Breakdown sheet) has a proper
hyperlink to https://tabby.atlassian.net/browse/<KEY> with blue underline styling.
"""
import openpyxl
from openpyxl.styles import Font

XLSX = "/Users/zaki/Downloads/Product discovery/Tech Debt/outputs/tech_debt_report.xlsx"
JIRA_BASE = "https://tabby.atlassian.net/browse/"

wb = openpyxl.load_workbook(XLSX)
ws = wb["Epic Breakdown"]

# Find Issue Key column (should be col 4 / D)
issue_key_col = None
header_row = None
for row in range(1, 4):
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row, col).value
        if val and "issue key" in str(val).lower():
            issue_key_col = col
            header_row = row
            break
    if issue_key_col:
        break

if not issue_key_col:
    print("ERROR: Could not find 'Issue Key' column header")
    exit(1)

print(f"Found 'Issue Key' at row {header_row}, col {issue_key_col}")

fixed = 0
skipped = 0
already_ok = 0

for row_idx in range(4, ws.max_row + 1):
    cell = ws.cell(row_idx, issue_key_col)
    val = cell.value
    if not val:
        skipped += 1
        continue

    key_str = str(val).strip()
    expected_url = JIRA_BASE + key_str

    has_hyperlink = (cell.hyperlink is not None) and bool(cell.hyperlink.target)
    hyperlink_correct = has_hyperlink and (cell.hyperlink.target == expected_url)

    if not hyperlink_correct:
        cell.hyperlink = expected_url
        cell.font = Font(
            name="Calibri",
            size=11,
            color="0563C1",   # standard Excel hyperlink blue
            underline="single",
        )
        fixed += 1
    else:
        already_ok += 1

wb.save(XLSX)
print(f"\nDone.")
print(f"  Fixed  : {fixed} cells")
print(f"  OK     : {already_ok} cells (already had correct hyperlink)")
print(f"  Skipped: {skipped} empty cells")
print(f"Saved: {XLSX}")
